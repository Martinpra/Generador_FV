import numpy as np
from openpyxl import load_workbook
from datetime import datetime

# Cargar datos desde el archivo Excel
archivo = load_workbook('Datos_climatologicos_Santa_Fe_2019.xlsx', read_only=True)
hoja = archivo.active

def cargar_datos_desde_excel():
    datos = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        datos.append(fila)
    return np.array(datos)

datos = cargar_datos_desde_excel()

def pot_modelo_GFV(G, T, N, Ppico, eta, kp, Pinv, mu=2, Gstd=1000, Tr=25):
    """
    Calcula la potencia generada por un GFV con los datos indicados.

    Parámetros:
    - G: Irradiancia en W/m² o instancia de datetime.datetime
    - T: Temperatura ambiente en °C
    - N: Cantidad de módulos
    - Ppico: Potencia pico de cada módulo en Watt
    - eta: Rendimiento global de la instalación
    - kp: Coeficiente de temperatura-potencia en °C^(-1)
    - Pinv: Potencia nominal del inversor en kilo-Watt
    - mu: Umbral porcentual mínimo del inversor
    - Gstd: Irradiancia estándar en W/m²
    - Tr: Temperatura de referencia en °C
    """

    # Verificar si G es una instancia de datetime.datetime
    if isinstance(G, datetime):
        # Obtener el valor numérico de G
        G = datos[G.minute + (G.hour - 1) * 6, 1]  # Índice en datos

    # Calcular temperatura de la celda (Tc) usando el modelo simple
    Tc = T + 0.031 * G

    # Calcular la potencia generada por cada módulo
    potencia_modulo = N * G / Gstd * Ppico * (1 + kp * (Tc - Tr)) * eta * 1e-3

    # Aplicar límites de generación y umbral mínimo del inversor
    potencia_generada = np.where(potencia_modulo > Pinv, Pinv, potencia_modulo)
    potencia_generada = np.where(potencia_generada < mu / 100 * Pinv, 0, potencia_generada)

    return potencia_generada

# Ejemplo de uso
instante = datetime(2019, 4, 24, 10, 20)  # Día 24 de abril a las 10:20
irradiancia, temperatura = datos[instante.minute + (instante.hour - 1) * 6, 1:3]  # Índice en datos
potencia_generada = pot_modelo_GFV(irradiancia, temperatura, 12, 240, 0.97, -0.0044, 2.5)

print(f"Instante: {instante}, Irradiancia: {irradiancia}, Temperatura: {temperatura}")
print(f"Potencia generada: {potencia_generada} kW")
